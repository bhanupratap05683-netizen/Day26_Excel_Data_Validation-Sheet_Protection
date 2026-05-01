# Day26_validation_protection.py
# Topic: Excel Data Validation & Sheet Protection using openpyxl
# File: Day26_Practice.xlsx (your practice file)

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

# ── STEP 1: Load the workbook ──────────────────────────────────────────────────
wb = load_workbook("Day26_Practice.xlsx")


# ═══════════════════════════════════════════════════════════════════
# PART A: DATA VALIDATION on Loan_Application sheet
# ═══════════════════════════════════════════════════════════════════
ws = wb["Loan_Application"]


# ── Validation 1: Department dropdown (C6) ────────────────────────
# type="list" means we give a comma-separated list of allowed values
dv_dept = DataValidation(
    type="list",
    formula1='"Finance,HR,Operations,IT,Marketing,Legal"',  # must be a string with quotes inside
    allow_blank=True,
    showDropDown=False,   # False = show the dropdown arrow to the user
    showErrorMessage=True,
    errorTitle="Invalid Department",
    error="Please select a valid department from the dropdown.",
    errorStyle="stop"     # "stop" blocks invalid input | "warning" allows but warns | "information" just informs
)
dv_dept.add("C6")        # attach to the Department input cell
ws.add_data_validation(dv_dept)


# ── Validation 2: Employment Type dropdown (C7) ───────────────────
dv_emp = DataValidation(
    type="list",
    formula1='"Full-Time,Part-Time,Contract,Intern"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Type",
    error="Select from: Full-Time, Part-Time, Contract, Intern",
    errorStyle="stop"
)
dv_emp.add("C7")
ws.add_data_validation(dv_emp)


# ── Validation 3: Annual Income — must be a whole number > 0 ─────
dv_income = DataValidation(
    type="whole",        # only integers allowed
    operator="greaterThan",
    formula1="0",
    showErrorMessage=True,
    errorTitle="Invalid Income",
    error="Annual income must be a positive number.",
    errorStyle="stop"
)
dv_income.add("C8")
ws.add_data_validation(dv_income)


# ── Validation 4: Credit Score — whole number between 300 and 900 ─
dv_credit = DataValidation(
    type="whole",
    operator="between",
    formula1="300",
    formula2="900",
    showInputMessage=True,     # shows a tooltip when user clicks the cell
    promptTitle="Credit Score",
    prompt="Enter a score between 300 and 900.",
    showErrorMessage=True,
    errorTitle="Invalid Credit Score",
    error="Credit score must be between 300 and 900.",
    errorStyle="stop"
)
dv_credit.add("C9")
ws.add_data_validation(dv_credit)


# ── Validation 5: Loan Type dropdown (C12) ────────────────────────
dv_loan = DataValidation(
    type="list",
    formula1='"Home Loan,Personal Loan,Education Loan,Vehicle Loan,Business Loan"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Loan Type",
    error="Please select a valid loan type.",
    errorStyle="stop"
)
dv_loan.add("C12")
ws.add_data_validation(dv_loan)


# ── Validation 6: Loan Amount — must be between 10,000 and 50,00,000 ──
dv_amount = DataValidation(
    type="whole",
    operator="between",
    formula1="10000",
    formula2="5000000",
    showInputMessage=True,
    promptTitle="Loan Amount",
    prompt="Enter amount between ₹10,000 and ₹50,00,000.",
    showErrorMessage=True,
    errorTitle="Invalid Amount",
    error="Loan amount must be between ₹10,000 and ₹50,00,000.",
    errorStyle="stop"
)
dv_amount.add("C13")
ws.add_data_validation(dv_amount)


# ── Validation 7: Tenure — between 6 and 360 months ──────────────
dv_tenure = DataValidation(
    type="whole",
    operator="between",
    formula1="6",
    formula2="360",
    showInputMessage=True,
    promptTitle="Tenure",
    prompt="Enter loan tenure in months (6 to 360).",
    showErrorMessage=True,
    errorTitle="Invalid Tenure",
    error="Tenure must be between 6 and 360 months.",
    errorStyle="stop"
)
dv_tenure.add("C14")
ws.add_data_validation(dv_tenure)


# ── Validation 8: Interest Rate — decimal between 1.0 and 30.0 ───
dv_rate = DataValidation(
    type="decimal",
    operator="between",
    formula1="1.0",
    formula2="30.0",
    showErrorMessage=True,
    errorTitle="Invalid Rate",
    error="Interest rate must be between 1.0% and 30.0%.",
    errorStyle="warning"  # warning: user can override if they really want
)
dv_rate.add("C15")
ws.add_data_validation(dv_rate)


# ═══════════════════════════════════════════════════════════════════
# PART B: DATA VALIDATION on Expenses_Raw sheet
# ═══════════════════════════════════════════════════════════════════
ws2 = wb["Expenses_Raw"]

# Dropdown for Department column (C3:C12)
dv_dept2 = DataValidation(
    type="list",
    formula1='"Finance,HR,Operations,IT,Marketing,Legal"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Department",
    error="Select a valid department.",
    errorStyle="stop"
)
dv_dept2.sqref = "C3:C12"   # apply to entire column range at once
ws2.add_data_validation(dv_dept2)

# Dropdown for Category column (D3:D12)
dv_cat = DataValidation(
    type="list",
    formula1='"Travel,Food,Software,Hardware,Training,Utilities,Other"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Category",
    error="Select a valid expense category.",
    errorStyle="stop"
)
dv_cat.sqref = "D3:D12"
ws2.add_data_validation(dv_cat)

# Amount must be positive (E3:E12)
dv_amt = DataValidation(
    type="decimal",
    operator="greaterThan",
    formula1="0",
    showErrorMessage=True,
    errorTitle="Invalid Amount",
    error="Amount must be greater than 0.",
    errorStyle="stop"
)
dv_amt.sqref = "E3:E12"
ws2.add_data_validation(dv_amt)

# Approved? must be Yes or No only (F3:F12)
dv_approved = DataValidation(
    type="list",
    formula1='"Yes,No,Pending"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Value",
    error="Approved field must be: Yes, No, or Pending.",
    errorStyle="stop"
)
dv_approved.sqref = "F3:F12"
ws2.add_data_validation(dv_approved)


# ═══════════════════════════════════════════════════════════════════
# PART C: SHEET PROTECTION on Loan_Application
# ═══════════════════════════════════════════════════════════════════
ws = wb["Loan_Application"]

# STEP 1: Unlock ALL cells first (openpyxl locks everything by default when protection is on)
for row in ws.iter_rows():
    for cell in row:
        cell.protection = Protection(locked=False)

# STEP 2: Lock the formula cells (Section C rows 19-23, column E)
formula_cells = ["E19", "E20", "E21", "E22", "E23"]
for cell_addr in formula_cells:
    ws[cell_addr].protection = Protection(locked=True)

# STEP 3: Enable sheet protection
# password is optional — without it, anyone can unprotect via Excel UI
# In real work, always set a password
ws.protection.sheet = True
ws.protection.password = "finance2026"   # your protection password
ws.protection.selectLockedCells = False  # user CAN click on locked cells (just can't edit)
ws.protection.selectUnlockedCells = False

print("✅ Sheet protection enabled on Loan_Application")
print("   Password: finance2026")
print("   Locked cells: E19, E20, E21, E22, E23 (formula cells)")
print("   All other cells: freely editable")


# ── SAVE ──────────────────────────────────────────────────────────
wb.save("Day26_Output.xlsx")
print("\n✅ Saved: Day26_Output.xlsx")
print("   - Loan_Application: 8 validation rules + formula protection")
print("   - Expenses_Raw: 4 validation rules on columns C, D, E, F")